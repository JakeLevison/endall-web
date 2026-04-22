#!/usr/bin/env python3
"""Render an xlsx as one HTML page per sheet for screenshot evidence.

Each sheet becomes its own HTML page: A1 is the title, headers in a dark
band, currency cells with $ prefix, formula cells shown as their formula
string (truncated) with the full formula in a title tooltip. The point of
the render is to prove (a) which cells hold user inputs and (b) which
cells hold real formulas referring to those inputs — together evidence
that the workbook is dynamic, not a static artifact.
"""

import sys
import html
from pathlib import Path

import openpyxl


def render(xlsx_path: Path, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    sheets = list(wb.sheetnames)

    # Wide tabs repeat the same 5-col block across 12 months. Crop to the
    # first N cols so formulas stay legible in a screenshot.
    MAX_COLS = {"Cash Flow": 11, "Sensitivity": 12}
    MAX_ROWS = {"Cash Flow": 30}

    for sn in sheets:
        ws = wb[sn]
        max_col = MAX_COLS.get(sn, ws.max_column or 1)
        max_row_count = MAX_ROWS.get(sn, ws.max_row or 1)
        rows_html = []
        for row in ws.iter_rows(max_col=max_col, max_row=max_row_count):
            cells = []
            for c in row:
                v = c.value
                if v is None:
                    cells.append('<td class="empty"></td>')
                elif isinstance(v, str) and v.startswith("="):
                    full = v
                    disp = v if len(v) <= 28 else v[:28] + "…"
                    cells.append(
                        f'<td class="formula" title="{html.escape(full)}">{html.escape(disp)}</td>'
                    )
                elif isinstance(v, (int, float)):
                    fmt = (c.number_format or "").lower()
                    if "$" in c.number_format or "dollar" in fmt:
                        cells.append(f'<td class="currency">${v:,.0f}</td>')
                    elif "%" in c.number_format:
                        cells.append(f'<td class="percent">{v * 100:.1f}%</td>')
                    else:
                        disp = f"{v:,}" if v else str(v)
                        cells.append(f'<td class="number">{html.escape(disp)}</td>')
                else:
                    cells.append(f'<td class="text">{html.escape(str(v))}</td>')
            rows_html.append("<tr>" + "".join(cells) + "</tr>")

        name = sn.replace(" ", "_")
        tabs = "".join(
            f'<span class="tab{" active" if t == sn else ""}">{html.escape(t)}</span>'
            for t in sheets
        )
        page = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>{html.escape(xlsx_path.name)} — {html.escape(sn)}</title>
<style>
body{{font:13px/1.3 -apple-system, Segoe UI, sans-serif; background:#f5f5f5; margin:0; padding:20px; color:#1f2937; display:inline-block;}}
.title{{font-size:14px; color:#555; margin-bottom:6px;}}
.filename{{font-family:monospace; color:#1f4e79;}}
.tabs{{display:flex; gap:2px; margin-bottom:0;}}
.tab{{display:inline-block; padding:6px 14px; background:#d6e4f0; border:1px solid #b4c6e7; border-bottom:none; font-size:12px; color:#444;}}
.tab.active{{background:#fff; color:#1f4e79; font-weight:600; border-top:2px solid #1f4e79;}}
table{{border-collapse:collapse; background:#fff; border:1px solid #b4c6e7; margin-top:0;}}
td{{border:1px solid #eaeaea; padding:4px 8px; min-width:50px; max-width:220px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; vertical-align:top;}}
td.empty{{color:#aaa;}}
td.text{{color:#1f2937;}}
td.number{{color:#1f2937; text-align:right;}}
td.currency{{color:#1f4e79; text-align:right; font-weight:600;}}
td.percent{{color:#1f4e79; text-align:right;}}
td.formula{{color:#6b21a8; font-family:monospace; font-size:11px;}}
.legend{{margin-top:14px; font-size:11px; color:#666;}}
.legend span{{display:inline-block; padding:2px 6px; border-radius:3px; margin-right:8px;}}
.legend .c{{background:#e0e7ff; color:#1f4e79;}}
.legend .f{{background:#f3e8ff; color:#6b21a8;}}
</style></head><body>
<div class="title">Rendered from <span class="filename">{html.escape(xlsx_path.name)}</span></div>
<div class="tabs">{tabs}</div>
<table>{"".join(rows_html)}</table>
<div class="legend">
  <span class="c">currency cell (static value)</span>
  <span class="f">formula cell (computes from references)</span>
</div>
</body></html>"""
        (out_dir / f"{xlsx_path.stem}__{name}.html").write_text(page)


if __name__ == "__main__":
    for xlsx in sys.argv[1:]:
        p = Path(xlsx)
        render(p, Path("/tmp/xlsx-render"))
        print("rendered", p)
